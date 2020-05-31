#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep 19 15:50:28 2018

@author: kaushal2612
"""

from openpyxl import load_workbook
from openpyxl import Workbook
import os.path
import re
import string

authorlist = os.listdir("author")
#print(authorlist)

for nameofauthor in range(0,len(authorlist)):
    xlsxfile= authorlist[nameofauthor]
    input_file = "author/"+xlsxfile
    wb = load_workbook(filename = input_file)
    print(nameofauthor, xlsxfile)
    # grab the active worksheet
    sheet = wb.worksheets[0]
    
    colindex =['A', 'B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
               'AA', 'AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']
    
    header = ['PMID',   'LAST_NAME',    'FIRST_NAME',   'MIDDLE_NAME',  'FIRST_INIT',   
              'SUFFIX', 'TITLE',        'AFFILIATION',  'EMAIL',        'CO-AUTHER',    
              'ABSTRACT','MESH_TERM',      'JOURNAL_TITLE','YEAR',         'LANG',    'ORCID','ID']
    for i in range(2,sheet.max_row+1):
        #pmid
        pmid = sheet.cell(row = i, column = 38).value
        if(pmid==None):
            pmid = ""

        #orcid 
        orcidlist = sheet.cell(row = i, column = 39).value
        orcid = '' 
        if(orcidlist!=None):
            orcidlist = orcidlist.split('||')
        #print(orcidlist)
        
        if(orcidlist!=None and len(orcidlist)>0):
            orcid = orcidlist[0]
        
        if not orcid:
            orcid = "No"
        
        authortext = sheet.cell(row = i, column = 3).value
        names = authortext.split("||")
        translator = str.maketrans('', '', string.punctuation)
        firstauthor = names[0].translate(translator)
        fullname = firstauthor.split(' ')
        #print(fullname)
        
        #lastname
        lastname = fullname[0]
        
        #firstname
        firstname = ""
        try:
            firstname = fullname[1]
        except:
            A = "FirstName is not present"

        #Lastname_FirstInitial
        lname_finitial=""
        try:
            lname_finitial+=fullname[1][0]
        except:
            q="First Initial is not there i.e only last name is present"
        try:
            lname_finitial+=fullname[2][0]
        except:
            q="First Initial is of 1 letter"
        
        #middlename
        middlename=""
        try:
            middlename = fullname[3]
        except:
            q="no middle name"
        
        #suffix
        
        #title
        title = str(sheet.cell(row = i, column = 30).value)
        
        #affiliation
        affiliationtext = sheet.cell(row = i, column = 41).value
        
        affi = ''
        if(authortext!=None and len(authortext)!=0):
            affiliation =[]
            if(affiliationtext!=None and len(affiliationtext)!=0 ):
                affiliation = affiliationtext.split("||")
            try:
                affi = affiliation[0]    
            except:
                A = "Affiliation is not present"
        affi = re.sub("\\n",'',affi)
        affi = re.sub("<affiliationinfo>",'',affi)
        affi = re.sub("</affiliationinfo>",'',affi)
        affi = re.sub("</affiliation>",'',affi)
        affi = re.sub("<affiliation>",'',affi)
        
        #email
        email=""
        if(affi!=''):
            affitext = affi.split(' ')
            isemail = affitext[len(affitext)-1]
            if(re.search('[^@]+@[^@]+\.[^@]+',isemail)):
                email=isemail
        
        #coauthor
        coauthor = []
        for ca in range(1,len(names)):
            a=names[ca]+"||"
            coauthor.append(a)
        

        #abstract
        abstract = sheet.cell(row = i, column = 8).value
        if abstract==None:
            abstract=""
        
        #mesh term
        meshterm = sheet.cell(row = i ,column = 40 ).value
        if meshterm == None:
            meshterm=""
        
        #journal
        journalcell = sheet.cell(row = i ,column = 15 ).value
        journal=""
        try:
            journaltext = journalcell.split('||')
            journalstring = journaltext[0]
            journal = re.findall(r'{"journal":"(.*)"}',journalstring)
        except:
            a = "Journal is empty"
        
        #year 
        year = ""
        yearinfo = sheet.cell(row = i ,column = 19 ).value
        try:
            year = yearinfo[0:4]
        except:
            a = "Year is empty"
        
        #lang
        lang = sheet.cell(row = i ,column = 18 ).value
        
        #id
        ndliid = str(sheet.cell(row = i ,column = 1 ).value)
        """
        print("ORCID", orcid)
        print("LASTNAME" , lastname)
        print("FIRSTNAME" , firstname)
        print("MIDDLENAME ",middlename)
        print("FirstInitial", lname_finitial)
        print("title", title)
        print("EMAIL",email)
        print("Co-author", coauthor)
        print("Abstract", abstract)
        print("meshterm", meshterm)
        print("journal", journal)
        print("Lang", lang)
        print("Year", year)
        """
        authordata = []
        authordata.append(pmid)
        authordata.append(lastname)
        authordata.append(firstname)
        authordata.append(middlename)
        authordata.append(lname_finitial)
        authordata.append("")   #suffix
        authordata.append(title)
        authordata.append(affi)
        authordata.append(email)
        authordata.append(coauthor)
        authordata.append(abstract)
        authordata.append(meshterm)
        authordata.append(journal)
        authordata.append(year)
        authordata.append(lang)
        authordata.append(orcid)
        authordata.append(ndliid)
        
        filepath = "author_features/"+xlsxfile
        if os.path.isfile(filepath)==False:
            filename = xlsxfile
            wb1 = Workbook()
            #add header
            ws1 = wb1.active
            
            wb1.save(filepath)
            
        wb1 = load_workbook(filename = filepath)
        
        ws1 = wb1.active
        sheetout = wb1.worksheets[0]
        
        #print(ndliid, rowno)
        for p in range(0,len(authordata)):
            ws1[colindex[p]+str(i-1)]=("".join(authordata[p]))
        wb1.save(filepath)
        