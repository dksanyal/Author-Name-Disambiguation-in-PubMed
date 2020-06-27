from openpyxl import load_workbook
from openpyxl import Workbook
import os

print("This program may take time depending upon the no of record in xlsx file inside Similarity_profile directory")
allfiles = os.listdir('Similarity_profile')
print("Printing all files in the above directory\n",allfiles,"\n")

target_filename = "Train_dataset.xlsx"
author=0
wb1 = Workbook()
#add header
ws1 = wb1.active

wb1.save(target_filename)


wbtrain = load_workbook(target_filename)
sheettrain = wbtrain.worksheets[0]

#6 7 3 4 5 1 2 1
header = ["field_1st_author","field_2nd_author","author_fname",    "author_midname", "auth_suffix", "author_lname_IDF",
           "affl_email","affl_jaccard", "affl_tfidf",  "affl_softtfidf", "affl_dept_jaccard", "affl_org_jaccard","affl_location_jaccard",
           "coauth_lname_shared",    "coauth_lname_idf",    "coauth_jaccard", "coauth_lname_finitial_jaccard",
           "mesh_shared", "mesh_shared_idf",    "mesh_tree_shared", "mesh_tree_shared_idf",
           "journal_shared_idf", "journal_year", "journal_year_diff",
           "abstract_jaccard",
           "title_jaccard","title_bigram_jaccard", "title_embedding_cosine", "abstract_embedding_cosine", "target"]
#adding header 
for i in range(0,len(header)):
    sheettrain.cell(row = 1, column = i+1).value = header[i]

trainrow = 2
for author in range(0,len(allfiles)):
    wbauthor = load_workbook("Similarity_profile/"+allfiles[author])
    sheetauthor = wbauthor.worksheets[0]
    if(sheetauthor.cell(row = 1, column =1).value != None):
        print("Extracting data from "+allfiles[author]+ ". This file contain total "+ str(sheetauthor.max_row+1) + " data")
        authorrow = 1
        for authorrow in range(1,sheetauthor.max_row+1):
            for authorcol in range(1,sheetauthor.max_column+1):
                sheettrain.cell(row = trainrow, column = authorcol).value = sheetauthor.cell(row = authorrow, column = authorcol).value
            trainrow = trainrow + 1

wbtrain.save(target_filename)

print("Created XLSX File now Generating CSV File")
import pandas as pd
data_xls = pd.read_excel(target_filename, 'Sheet', index_col=None)
data_xls.to_csv('Train_dataset.csv', encoding='utf-8', index=False)