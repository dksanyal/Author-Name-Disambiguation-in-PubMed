import os
from pyjarowinkler import distance
import math
import re
import xlsxwriter
import string
from xlrd import open_workbook
from collections import Counter
from bs4 import BeautifulSoup
from nltk import bigrams
from openpyxl import load_workbook
from numpy import dot
from numpy.linalg import norm
from flair.data import Sentence
from flair.embeddings import BertEmbeddings
from gensim.models import Word2Vec
from flair.embeddings import WordEmbeddings, DocumentPoolEmbeddings, FlairEmbeddings

meshxml = BeautifulSoup(open("2018MeShTreeHierarchy.xml"),'lxml').get_text()

authorlist = os.listdir("author_features")

stopwords = ["a","about","again","all","almost","also","although","always","among","an","and","another","any","are","as","at"
            ,"be","because","been","before","being","between","both","but","by"
            ,"can","could"
            ,"did","do","does","done","due","during"
            ,"each","either","enough","epecially","etc"
            ,"for","found","from","further"
            ,"had","has","have","having","here","how","however"
            ,"i","if","in","into","is","its","itself"
            ,"just"
            ,"kg","km"
            ,"made","mainly","make","may","mg","might","ml","mm","most","mostly","must"
            ,"nearly","neither","no","nor"
            ,"obtained","of","often","on","our","overall"
            ,"perhaps","pmid"
            ,"quite"
            ,"rather","really","regarding"
            ,"seem","seen","several","should","show","showed","shown","shows","significantly","since","so","some","such"
            ,"than","that","the","their","theirs","them","then","there","therefore","these","they","this","those","through","thus","to"
            ,"upon","use","used","using"
            ,"various","very",",","."
            ,"was","we","were","what","when","which","while","with","within","without","would"]

def removestopword(words):
    new_words = [word for word in words if word not in stopwords]
    return new_words

result = []

def common_elements(list1, list2):
    common = []
    for i in range(0,len(list1)):
        for j in range(0,len(list2)):
            if(list1[i]==list2[j]):
                common.append(list1[i])
                break
    return common

def auth_fst(authorA, authorB):
    if(authorA[2]=='' or authorB[2]==''):
        return 0
    
    if ((authorA[2] == authorB[2])):
        return 3
    if((authorA[2] != authorB[2])):
        return 2
    return 0
    
def auth_mid(midA, midB):
    if(midA=='' or midB=='' ):
        return 0
    if(midA == midB):
        return 3
    if(midA != midB):
        return 2
    return 0
    
def auth_suf(sufA, sufB):
    if(sufA==sufB and sufA!='' and sufB!=''):
        return 1
    else:
        return 0

print("Loading Lastname Frequency")
wblname = load_workbook(filename="Metadata_frequency/lname_frequency.xlsx")
sheetlname = wblname.worksheets[0]
klname = dict()
for i in range(1,sheetlname.max_row+1):
    klname[str(sheetlname.cell(row= i, column=1).value).lower()] = sheetlname.cell(row= i, column=2).value
    
def auth_lname_idf(lnameA, lnameB):
    if (lnameA != lnameB):
        return 0
    else:
        #print(type(klname.get('totallname')))
        lnameidf = int(klname.get('totallname'))/klname.get(lnameA)
        return math.log( lnameidf )

#used in aff_tfidf author similarity
def idf(t):
    total = 0
    for i in range(0,len(data)):
        translator = str.maketrans('', '', string.punctuation)
        aff = data[i][7].lower()
        aff = aff.strip()
        aff = aff.replace(data[i][8], '')
        aff = aff.strip(',')
        aff = aff.translate(translator)    
        
        if(aff.find(t) != -1 ):
            total += 1
    return (len(data))/total
    
def aff_email(emailA,emailB):
    if(emailA!='' and emailB!='' and emailA==emailB):
        return 1
    return 0

def aff_jac(affA, affB, emailA, emailB):
    affA = affA.lower()
    affB = affB.lower()
    affA = affA.replace(emailA, '')
    affB = affB.replace(emailB, '')
    
    if (affA=='' or affB==''):
        return 0
    affA = affA.split()
    affB = affB.split()
    affA = removestopword(affA)
    affB = removestopword(affB)
    common = len(common_elements(affA,affB))
    return common/(len(affA)+len(affB))

print("Loading Affiliation Frequency")
wbaff = load_workbook(filename="Metadata_frequency/aff_frequency.xlsx")
sheetaff = wbaff.worksheets[0]
kaff = dict()
for i in range(1,sheetaff.max_row+1):
    kaff[str(sheetaff.cell(row= i, column=1).value).lower()] = sheetaff.cell(row= i, column=2).value


def aff_tfidf(affA, affB, emailA, emailB):
    affA = affA.lower()
    affB = affB.lower()
    
    affA = affA.replace(emailA, '')
    affB = affB.replace(emailB, '')
    if (affA=='' or affB==''):
        return 0
    
    translator = str.maketrans('', '', string.punctuation)
    affA = affA.strip(',')
    affA = affA.translate(translator)    
    affB = affB.strip(',')
    affB = affB.translate(translator)    
    
    affA = affA.split()
    affB = affB.split()
    affA = removestopword(affA)
    affB = removestopword(affB)
    common = common_elements(affA,affB)
    
    tfA = Counter(affA)
    
    tfB = Counter(affB)
        
    tfc = Counter(common)
    
    tfidfA= [];
    tfidfB = []
    for x,y in tfc.items():
        try:
            affidfx = int(kaff.get('totalaff'))/kaff.get(x)
        except:
            affidfx = int(kaff.get('totalaff'))/1
        tfidfA.append( (math.log(tfA[x] + 1)) * math.log( affidfx ) )
    for x,y in tfc.items():
        try:
            affidfx = int(kaff.get('totalaff'))/kaff.get(x)
        except:
            affidfx = int(kaff.get('totalaff'))/1
        tfidfB.append( (math.log(tfB[x] + 1)) * math.log ( affidfx ) )
    
    sumc = 0
    
    for i in range(0,len(tfidfA)):
        sumc += tfidfA[i]*tfidfB[i]
    #print(sumc)
    
    return sumc 

#tfidf for aff_softtfidf TFIDF(t,affA)

def tfidf(t,aff):
    tf = Counter()
    for doc in aff:
        for word in doc.split():
            tf[word] +=1
    try:
        return ( math.log( tf[t] + 1 ) * math.log( int(kaff.get('totalaff')) /kaff.get(t) ) )
    except:
        return 0

def aff_softtfidf(affA, affB, emailA, emailB):
    affA = affA.lower()
    affB = affB.lower()
    
    affA = affA.replace(emailA, '')
    affB = affB.replace(emailB, '')
    if (affA=='' or affB==''):
        return 0
    
    translator = str.maketrans('', '', string.punctuation)
    affA = affA.strip(',')
    affA = affA.translate(translator)    
    affB = affB.strip(',')
    affB = affB.translate(translator)    
    
    affA = affA.split()
    affB = affB.split()
    affA = removestopword(affA)
    affB = removestopword(affB)
    
    N = []
    for t in range (0,len(affA)):
        max1 = -999999
        localmax = max1
        for v in range (0, len(affB)):
            jarovalue = distance.get_jaro_distance(affA[t],affB[v])
            if(jarovalue < 0.8):
                localmax = jarovalue
            if(localmax>max1):
                max1=localmax
        N.append(max1)
    
    sum1 = 0
    for t in range(0, len(affA)):
        sum1+= tfidf( affA[t] ,affA ) * tfidf ( affA[t] ,affB ) * N[t]
    return sum1    

def aff_ner_jac(deptA, deptB, orgA, orgB, locA, locB):
    
    dept_jac = org_jac = loc_jac = 0
    if(deptA=="" or deptB=="" or deptA=="none" or deptB=="none"):
        dept_jac = 0
    else:
        deptA= deptA.split("||")
        deptB= deptB.split("||")
        deptA = removestopword(deptA)
        deptB = removestopword(deptB)
        deptcommon = len(common_elements(deptA,deptB))
        if deptcommon!=0:
            dept_jac = deptcommon/(len(deptA)+len(deptB))
        
    if(orgA=="" or orgB=="" or orgA=="none" or orgB=="none"):
        org_jac = 0
    else:
        orgA= orgA.split("||")
        orgB= orgB.split("||")
        orgA = removestopword(orgA)
        orgB = removestopword(orgB)
        orgcommon = len(common_elements(orgA,orgB))
        if orgcommon!=0:
            org_jac = orgcommon/(len(orgA)+len(orgB))
        
    if(locA=="" or locB=="" or locA=="none" or locB=="none"):
        loc_jac = 0
    else:
        locA= locA.split("||")
        locB= locB.split("||")
        locA = removestopword(locA)
        locB = removestopword(locB)
    
        loccommon = len(common_elements(locA,locB))
        if loccommon!=0:
            loc_jac = loccommon/(len(locA)+len(locB))
        
    return dept_jac, org_jac, loc_jac

def get_lname_author(coauthorA,coauthorB):
    lastnamesA=[]
    for i in range(0,len(coauthorA)):
        fullname = coauthorA[i].split(' ')
        #lastname
        lastnamesA.append(fullname[0])
    
    lastnamesB=[]
    for i in range(0,len(coauthorB)):
        fullname = coauthorB[i].split(' ')
        #lastname
        lastnamesB.append(fullname[0])
        
    common = list(common_elements(lastnamesA,lastnamesB))
    #print(common)
    return common

print("Loading Coauthor Lastname Frequency")
wbcoauth = load_workbook(filename="Metadata_frequency/coauthor_frequency.xlsx")
sheetcoauth = wbcoauth.worksheets[0]
kcoauth = dict()
for i in range(1,sheetcoauth.max_row+1):
    kcoauth[str(sheetcoauth.cell(row= i, column=1).value).lower()] = sheetcoauth.cell(row= i, column=2).value


def coauth_lname_shared(coauthorA,coauthorB):
    coauthorA = coauthorA.lower()
    coauthorB = coauthorB.lower()
    coauthorA = coauthorA.split('||')
    coauthorB = coauthorB.split('||')
    coauthorA.pop(len(coauthorA)-1)
    coauthorB.pop(len(coauthorB)-1)
        
    common = get_lname_author(coauthorA,coauthorB)
    #print(common)
    return len(common)

def coauth_lname_idf(coauthorA, coauthorB):

    coauthorA = coauthorA.split('||')
    coauthorB = coauthorB.split('||')
    coauthorA.pop(len(coauthorA)-1)
    coauthorB.pop(len(coauthorB)-1)
    common = list(common_elements(coauthorA,coauthorB))
    sum1=0
    # Extracting the last name of common author
    for i in range(0,len(common)):
        common[i] = common[i].split()[0]
    #print('check ', common)
    
    for i in range(0,len(common)):
        try:
            coauthidf = kcoauth.get('totalcoauth')/kcoauth.get(str(common[i]))
        except:
            coauthidf=1
        sum1 += math.log( coauthidf )
    
    return sum1
    
def coauth_lname_jac(coauthorA,coauthorB,comlen):
    if(coauthorA=='' or coauthorB==''):
        return 0
    coauthorA = coauthorA.split('||')
    coauthorB = coauthorB.split('||')
    coauthorA.pop(len(coauthorA)-1)
    coauthorB.pop(len(coauthorB)-1)
    return ( comlen / (len(coauthorA) + len(coauthorB) ) )

def coauth_lnamefi_jac(coauthorA,coauthorB):
    if(coauthorA=='' or coauthorB==''):
        return 0
    coauthorA = coauthorA.split('||')
    coauthorB = coauthorB.split('||')
    coauthorA.pop(len(coauthorA)-1)
    coauthorB.pop(len(coauthorB)-1)
    common = list(common_elements(coauthorA,coauthorB))
    return ( len(common) / (len(coauthorA) + len(coauthorB) ) )

def splitmesh(orimeshA):
    meshA=[]
    s=''
    orimeshA = re.sub(r'\*','',str(orimeshA))
    orimeshA = re.sub(r'\, ','',str(orimeshA))
    for i in range(0,len(orimeshA)):
        if(orimeshA[i]=='/' or (orimeshA[i]=='|' and orimeshA[i-1]=='|')):
            meshA.append(s)
            s=''
        else:
            s=s+orimeshA[i]
    
    meshA = [i.strip(',') for i in meshA]
    meshA = [i.strip('|') for i in meshA]
    meshA = [i.strip('*') for i in meshA]
    meshA = list(set(meshA))
    return meshA

print("Loading MeshTerm Frequency")
wbmesh = load_workbook(filename="Metadata_frequency/mesh_frequency.xlsx")
sheetmesh = wbmesh.worksheets[0]
kmesh = dict()
for i in range(1,sheetmesh.max_row+1):
    kmesh[str(sheetmesh.cell(row= i, column=1).value)] = sheetmesh.cell(row= i, column=2).value

def mesh_shared_idf(orimeshA, orimeshB):
    
    meshA=splitmesh(orimeshA)
    meshB=splitmesh(orimeshB)

    sum1=0
    common = list(common_elements(meshA,meshB))
    for i in range(0,len(common)):
        if(len(common[i].split(' '))<=1):
            common[i] = common[i].capitalize()
        try:
            #print(common[i],kmesh.get((common[i])))
            meshidf = kmesh.get('totalmesh')/kmesh.get(str(common[i]))
        except:
            meshidf=1
        sum1 += math.log( meshidf )
    
    return sum1

def mesh_tree_shared(orimeshA, orimeshB):
    meshA = []
    meshB = []
    s=''
    orimeshA = re.sub(r'\*','',str(orimeshA))
    orimeshB = re.sub(r'\*','',str(orimeshB))
    #orimeshA = re.sub(r'\, ','||',str(orimeshA))
    #orimeshB = re.sub(r'\, ','||',str(orimeshB))
    for i in range(0,len(orimeshA)):
        if(orimeshA[i]=='/' or (orimeshA[i]=='|' and orimeshA[i-1]=='|')):
            meshA.append(s)
            s=''
        else:
            s=s+orimeshA[i]
    
    meshA = [i.strip('|') for i in meshA]
    meshA = [i.strip('*') for i in meshA]
    meshA = list(set(meshA))
    
    s=''
    for i in range(0,len(orimeshB)):
        if(orimeshB[i]=='/' or (orimeshB[i]=='|' and orimeshB[i-1]=='|')):
            meshB.append(s)
            s=''
        else:
            s=s+orimeshB[i]
    meshB = [i.strip('|') for i in meshB]
    meshB = [i.strip('*') for i in meshB]
    meshB = list(set(meshB))
    
    pathmeshA = []
    pathmeshB = []
    temp = []
    for i in range(0,len(meshA)):
        try:
            temp = re.findall(r'[A-Z].*' + meshA[i] + '\n',meshxml)
            temp = [k.strip('\n') for k in temp]
            temp = [re.split("\.\.+", k) for k in temp]
        except:
            tempstr = ""
            word = meshA[i]
            for index in range(0,len(word)):
                if(word[index]=="+" or word[index]==")" or word[index]=="(" or word[index]=="-"):
                    tempstr = tempstr + "\\"
                tempstr = tempstr + word[index]
            #print(tempstr)
            temp = re.findall(r'[A-Z].*' + tempstr + '\n',meshxml)
            temp = [k.strip('\n') for k in temp]
            temp = [re.split("\.\.+", k) for k in temp]
        
        tree = []
        for eachind in range(0,len(temp)):
            if(len(temp)>0 and temp[eachind][1] == meshA[i]):
                #print(temp[eachind])  
                tree.append(temp[eachind])
        
        pathmeshA.append(tree)
    
    temp = []
    for i in range(0,len(meshB)):
        try:
            temp = re.findall(r'[A-Z].*' + meshB[i] + '\n',meshxml)
            temp = [k.strip('\n') for k in temp]
            temp = [re.split("\.\.+", k) for k in temp]
        except:
            tempstr = ""
            word = meshB[i]
            for index in range(0,len(word)):
                if(word[index]=="+" or word[index]==")" or word[index]=="(" or word[index]=="-"):
                    tempstr = tempstr + "\\"
                tempstr = tempstr + word[index]
            #print(tempstr)
            temp = re.findall(r'[A-Z].*' + tempstr + '\n',meshxml)
            temp = [k.strip('\n') for k in temp]
            temp = [re.split("\.\.+", k) for k in temp]
        
        tree = []
        for eachind in range(0,len(temp)):
            if(len(temp)>0 and temp[eachind][1] == meshB[i]):
                #print(temp[eachind])  
                tree.append(temp[eachind])
        pathmeshA.append(tree)
    

        pathmeshB.append(temp)
    
    nooftexttag = len( re.findall(r'[A-Z]\d.*\n',meshxml) )
    
    count = 0
    idfcount=0
    for i in range(0,len(pathmeshA)):
        for kA in range(0,len(pathmeshA[i])):
            treepathA = pathmeshA[i][kA][0]
            #print(treepathA)
            #treemeshtermA = pathmeshA[i][kA][1]
            while(treepathA !=''):
                matchA = 0
                for j in range(0,len(pathmeshB)):
                    if(matchA==1):
                        break
                    for kB in range(0,len(pathmeshB[j])):
                        if(matchA==1):
                            break
                        treepathB = pathmeshB[j][kB][0]
                        #treemeshtermB = pathmeshB[j][kB][1]
                        
                        while(treepathB != ''):
                            if(treepathB == treepathA ):
                                count+=1
                                #print('TreepathA ',treepathA, 'TreepathB ',treepathB)
                                matchA = 1
                                patternAB = treepathA + "\.\.\.\."
                                meshofpatternAB = (re.findall(patternAB+".*\n",meshxml))
                                temp = [k.strip('\n') for k in meshofpatternAB]
                                temp = [re.split("\.\.+", k) for k in temp]
                                #print("temp ", temp[0][1], kmesh.get(temp[0][1]))
                                # search in xml file
                                try:
                                    idfcount += math.log(kmesh['totalmesh']/kmesh.get(temp[0][1]))
                                except:
                                    # Mesh Term is rare
                                    idfcount += math.log(kmesh['totalmesh']/1)
                                    try:
                                        kmesh[temp[0][1]] = 1
                                    except:
                                        kmesh[temp[0][1]] = kmesh[temp[0][1]] + 1
                                #print(treepathA, ' \n', treepathB, ' ', noofterm, ' ', idfcount)
                                break
                            treepathB=treepathB[:-4]
                            
                treepathA=treepathA[:-4]
    #print(pathmeshB)
    return count,idfcount        
    
print("Loading Journal Frequency")
wbjour = load_workbook(filename="Metadata_frequency/journal_frequency.xlsx")
sheetjour = wbjour.worksheets[0]
kjour = dict()
for i in range(1,sheetjour.max_row+1):
    kjour[str(sheetjour.cell(row= i, column=1).value).lower()] = sheetjour.cell(row= i, column=2).value

def jour_shared_idf(jourA, jourB):
    if(jourA!=jourB):
        return 0
    elif(jourA=='' or jourB==''):
        return 0
    else:
        journalidf = int(kjour.get('jourtotal'))/kjour[jourA]
        return math.log( journalidf )

def jour_lang(langA, langB):
    if ( langA=='' or langB==''):
        return 0
    if ( langA == langB ):
        return 1
    elif ( langA != langB):
        return 2

print("Loading Language Frequency")
wblang = load_workbook(filename="Metadata_frequency/lang_frequency.xlsx")
sheetlang = wblang.worksheets[0]
klang = dict()
for i in range(1,sheetlang.max_row+1):
    klang[str(sheetlang.cell(row= i, column=1).value).lower()] = sheetlang.cell(row= i, column=2).value

def jour_lang_idf(langA,langB):
    #print(langA[14], type(langA[14]))
    llangA = langA[14]
    llangB = langB[14]
    if(llangA != llangB ):
        return 1
    elif(llangA=='' or llangB==''):
        return 0
    else:
        #print(klang.get('langtotal'), klang.get(llangA), llangA)
        langidf = int(klang.get('langtotal'))/klang.get(llangA)
        return math.log(langidf)

def jour_year(yearA, yearB):
    if(yearA=='' or yearB==''):
        return 5
    yearA = yearA.split()
    yearB = yearB.split()
    if(int(yearA[0]) < 1988 and int(yearB[0]) < 1988):
        return 0
    elif((int(yearA[0]) < 1988 and int(yearB[0]) >= 1988) or (int(yearA[0]) >= 1988 and int(yearB[0]) < 1988) ):
        return 1
    elif((int(yearA[0]) >= 1988 and int(yearB[0]) < 2002) or (int(yearA[0]) >= 1988 and int(yearB[0]) < 2002)):
        return 2
    elif(int(yearA[0]) >= 2002 and int(yearB[0]) >= 2002):
        return 4
    elif((int(yearA[0]) >= 1988 and int(yearB[0]) >= 2002) or (int(yearA[0]) >= 1988 and int(yearB[0]) >= 2002)):
        return 3

def jour_year_diff(yearA,yearB):
    if(yearA=='' or yearB==''):
        return 0
    yearA = yearA.split()
    yearB = yearB.split()
    return abs(int(yearA[0])-int(yearB[0]))    

def title_shared_jac(titleA,titleB):
    common = list(common_elements(titleA,titleB))
    return ( len(common) / (len(titleA)+len(titleB) ) )
    
def title_bigram_jac(titleA,titleB):
    if(len(titleA)==0 or len(titleB)==0):
        return 0
    bigramA = list(bigrams(titleA))
    bigramB = list(bigrams(titleB))
    common = common_elements(bigramA,bigramB)
    return ( len(common) / (len(titleA)+len(titleB) ) )
    
def author_similarity(author,similarity):
    #auth_fst = first name similarity
    fst = auth_fst(author[0],author[1])
    #auth_mid = mid name similarity
    mid = auth_mid(author[0][3],author[1][3])
    #auth_suf = suffix name similarity
    suf = auth_suf(author[0][5],author[1][5])
    #idf weight of the last name, IDF is the inverse of the fraction of names in the corpus.
    lname_idf = auth_lname_idf(author[0][1], author[1][1])
    
    similarity.append(fst)
    similarity.append(mid)
    similarity.append(suf)
    similarity.append(lname_idf)

    
def affiliation_similarity(author,similarity):
    email = aff_email(author[0][8],author[1][8])
    #jac = the jaccard similarity between affA and affB
    jac = aff_jac(author[0][7],author[1][7],author[0][8],author[1][8])
    #tfidf = the sum of TFIDF weights of shared terms in affA and affB
    tfidf1 = aff_tfidf(author[0][7],author[1][7],author[0][8],author[1][8])
    #softtfidf = the soft-TFIDF distance between affA
    #and affB . The soft-TFIDF distance is a hybrid distance
    #that combines a string-based distance with the TFIDF dis-
    #tance.
    softtfidf = aff_softtfidf(author[0][7],author[1][7],author[0][8],author[1][8])
    
    dept_jac, org_jac, loc_jac = aff_ner_jac(author[0][16],author[1][16],author[0][17],author[1][17],author[0][18],author[1][18])
    similarity.append(email)
    similarity.append(jac)
    similarity.append(tfidf1)
    similarity.append(softtfidf)
    similarity.append(dept_jac)
    similarity.append(org_jac)
    similarity.append(loc_jac)
    
    

def coauthor_similarity(author,similarity):
    #The number of shared coauthor last names between 2 paper
    lname_shared = coauth_lname_shared(author[0][9],author[1][9])
    #the sum of IDF values of all shared coauthor last name 
    lname_idf = coauth_lname_idf(author[0][9],author[1][9])
    #the jaccord similarity between couthA and coauthB
    lname_jac = coauth_lname_jac(author[0][9],author[1][9],lname_shared)
    lnamefi_jac = coauth_lnamefi_jac(author[0][9],author[1][9])
    
    similarity.append(lname_shared)
    similarity.append(lname_idf)
    similarity.append(lname_jac)
    similarity.append(lnamefi_jac)
    
    
def concept_similarity(author,similarity):
    #the number of shared mesh term b/n 2 paper same as Co-auth lname shared
    shared = coauth_lname_shared(author[0][11], author[1][11])
    #the sum of idf values of all shared mesh terms
    shared_idf = mesh_shared_idf(author[0][11], author[1][11])
    #Define a set T (mesh i ) to be the set of all ancestor concepts of mesh i in MeSH hierarchy.
    #tree_shared = mesh_tree_shared(author[0][11], author[1][11])
    tree_shared,tree_shared_idf = mesh_tree_shared(author[0][11], author[1][11])
    
    similarity.append(shared)
    similarity.append(shared_idf)
    similarity.append(tree_shared)
    similarity.append(tree_shared_idf)
    
def journal_similarity(author, similarity):
    #IDF value of shared journal, both are published in the same journal
    shared_idf = jour_shared_idf(author[0][12],author[1][12])
    #lang = journal language
    lang = jour_lang(author[0][14],author[1][14])
    #lang_idf = IDF value of shared journal language
    lang_idf = math.log( jour_lang_idf(author[0],author[1]) )
    #lang_year = categorical variable reflecting change in Med-line’s policy
    year = jour_year(author[0][13],author[1][13])
    #jour year diff
    year_diff = jour_year_diff(author[0][13],author[1][13])
    
    similarity.append(shared_idf)
    similarity.append(lang)
    similarity.append(lang_idf)
    similarity.append(year)
    similarity.append(year_diff)

def abstract_similarity(author,similarity):
    abstractA = author[0][10]
    abstractB = author[1][10]
    if(abstractA=='' or abstractB==''):
        return 0
    translator = str.maketrans('', '', string.punctuation)
    abstractA = abstractA.strip(',')
    abstractA = abstractA.translate(translator)    
    abstractB = abstractB.strip(',')
    abstractB = abstractB.translate(translator)    
    
    abstractA = abstractA.split()
    abstractB = abstractB.split()
    abstractA = removestopword(abstractA)
    abstractB = removestopword(abstractB)
    common = list(common_elements(abstractA,abstractB))
    return float( float(len(common)) / float(len(abstractA)+len(abstractB) ) )
    
def title_similarity(author, similarity):
    
    titleA = author[0][6]
    titleB = author[1][6]
    translator = str.maketrans('', '', string.punctuation)
    titleA = titleA.strip(',')
    titleA = titleA.translate(translator)    
    titleB = titleB.strip(',')
    titleB = titleB.translate(translator)    
    
    titleA = titleA.split()
    titleB = titleB.split()
    titleA = removestopword(titleA)
    titleB = removestopword(titleB)
    
    shared_jac = title_shared_jac(titleA,titleB)
    bigram_jac = title_bigram_jac(titleA,titleB)
    
    similarity.append(shared_jac)
    similarity.append(bigram_jac)

bert_embedding = BertEmbeddings('bert-base-uncased')
model = [bert_embedding]

def cosine_embedding(sentence1, sentence2, model):
    embeddings = DocumentPoolEmbeddings(model, mode='mean')
    s1 = Sentence(sentence1)
    s2 = Sentence(sentence2)
    e1 = embeddings.embed(s1)
    e2 = embeddings.embed(s2)
    v1 = s1.get_embedding()
    v2 = s2.get_embedding()
    #print(v1, v2)   #check that you don't get empty tensors

    cos_sim = dot(v1, v2)/(norm(v1)*norm(v2))
    #print(cos_sim)
    return cos_sim                       #lies in [-1, 1].
    
def title_abstract_embedding(author, similarity):
    titlea = author[0][6]
    titleb = author[1][6]
    abstracta = author[0][10]
    abstractb = author[1][10]
    
    title_embedding = abstract_embedding = 0.0
    if(titlea!='' or titleb!='' or titlea is not None or titleb is not None ):
        title_embedding = cosine_embedding(titlea, titleb, model)
    try:
        if(abstracta!='' or abstractb!='' or (type(abstracta)!=None and type(abstractb)!=None)):
            abstract_embedding = cosine_embedding(abstracta, abstractb, model)
    except:
        a="abstract is empty" 
        #print(a)
    similarity.append(title_embedding)
    similarity.append(abstract_embedding)        
    
def main(authorname):
    excel_file = xlsxwriter.Workbook('Similarity_profile/'+authorname)
    sheet = excel_file.add_worksheet('data')

    row = 0
    col = 0
    i=0
    save = 0
    if i==0:
        for i in range(0,len(data)):
            j=i+1
            for j in range(i+1,len(data)):
                # for tracking purpose
                print(i+1 ,' ', data[i][0], ' ', j+1 , ' ' , data[j][0], ' ' , data[i][1] + " " + data[i][4], len(data))
                author = []
                author.append(data[i])
                author.append(data[j])
                
                similarity = []
                similarity.append(str(data[i][1] + " " + data[i][4] + "_" + data[i][15]))
                similarity.append(str(data[j][1] + " " + data[j][4] + "_" + data[j][15]))
                
                #author similarity
                #author_similarity(author,similarity)
                #affiliation similarity
                #affiliation_similarity(author,similarity)
                #coauthor similarity
                coauthor_similarity(author,similarity)
                #concept similarity with the help of mesh term
                #concept_similarity(author,similarity)
                #journal similarity
                #journal_similarity(author, similarity)
                #abstract similarity
                #abstract_jac = abstract_similarity(author,similarity)
                #similarity.append(abstract_jac)
                #title similarity
                #title_similarity(author, similarity)
                #title_abstract_embedding(author, similarity)
                # Result whether those two are same or not
                same=0
                orcidcell1 = data[i][15]
                orcidcell1 = orcidcell1.split('||')
                orcid1 = orcidcell1[0]
                
                orcidcell2 = data[j][15]
                orcidcell2 = orcidcell2.split('||')
                orcid2 = orcidcell2[0]
                
                same=0
                #print(orcid1 , orcid2 , orcid1==orcid2)
                if(orcid1=="no" or orcid2=="no"):
                    same=0
                # if anyone of them is matches
                # orcid is there in two form XXXX-XXXX-XXXX-XXXX or http://orcid.org/XXXX-XXXX-XXXX-XXXX
                elif(orcid1==orcid2 or orcid1=='http://orcid.org/'+orcid2 or orcid2=='http://orcid.org/'+orcid1 ):
                    same=1
                similarity.append(same)
                    
                result.append(similarity)
                col = 0
                for col_value in range(0,len(similarity)):
                    sheet.write(row,col,similarity[col_value])
                    col += 1
                row += 1
                j+=1
            
    if save == 0:
        excel_file.close()


if __name__ == "__main__":
    print("Total No of Author: ",len(authorlist))
    print("\nFor Tracking: Author1_Row_in_Excel, Author1_PMID, Author2_Row_in_Excel,  Author2_PMID, Author_Lname_FirstInitial, No of Record are printed")
    for author in range(0,len(authorlist)):
        
        global rb 
        print(author, "author_features/"+authorlist[author])
        rb = open_workbook("author_features/"+authorlist[author])
        global sheet
        sheet = rb.sheet_by_name('Sheet')
        global data
        data = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        
        for i in range(0,len(data)):
            for j in range(0,len(data[0])):
                if(type(data[i][j])!=type("sgskh")):
                    r=0
                else:
                    if(j==11):
                        affmesh = "Meshterm cannot be lowered"
                    else:    
                        data[i][j]=data[i][j].lower()
        
        main(authorlist[author])    